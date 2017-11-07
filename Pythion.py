#!/usr/bin/python
# -*- coding: utf8 -*-

import numpy as np # arrays and mathematics
import scipy #scientific
import pandas as pd #data manipulation and analisys
import sys  #system-specific parameters and functions
import h5py #for huge amounts of numerical data in binaries
import os #pathname manipulations
import time #time access and conversions
import platform #what system we have


from collections import OrderedDict #for making ordered dictionaries
from openpyxl import Workbook #creation Exel file
from openpyxl import load_workbook #loading exel from existing file

#some stuff from scipy

from scipy import io as spio
from scipy import ndimage

#some stuff from pandas
import pandas.io.parsers

#graph drawning modules
#from plotgui4k import *
#from plotguiretina import *
import pyqtgraph as pg
import pyqtgraph.exporters
from pyqtgraph.dockarea import *





########written by developers########
import UsefulFunctions as uf #module for data processing
import Load #module for data processing
from UserInterface import * #loading pyqt5 designer generated graphical visualisation for program
from abfheader import * #code for proccessing some binaries
from CUSUMV2 import detect_cusum #CUSUM algorithm
from CUSUM import * #CUSUM algorithm
from PoreSizer import * #poresizer
from batchinfo import * #for visuallization
from MergeChimera import load_and_merge # Load and merge Chimera files
from Cut import *
from matplotlib.backends.backend_pdf import PdfPages #PDF pages creation
from Plot import * 
########Starting the program########
class GUIForm(QtGui.QMainWindow):


    def __init__(self, master=None):
        ####Setup GUI and draw elements from UI file#########
        QtGui.QMainWindow.__init__(self,master)
        self.ui = Ui_PythIon()
        self.ui.setupUi(self)

        ##########Linking buttons to main functions############

        ##########CUT LOAD BASELINE etc. BUTTONS connection to implementation functions##################################
        self.ui.loadbutton.clicked.connect(self.loadfile) #loadbutton
        self.ui.clearscatterbutton.clicked.connect(self.clearscatter) #clear data
        self.ui.analyzebutton.clicked.connect(self.analyze) #analyze using Low Pass filter
        #self.ui.baselinebutton.clicked.connect(self.baselinecalc) #baseline
        self.ui.fitbutton.clicked.connect(self.Analyze) #fit using CUSUM
        self.ui.cutbutton.clicked.connect(self.CutData) #cutbutton
        self.ui.nextfilebutton.clicked.connect(self.nextfile) #NEXT
        self.ui.previousfilebutton.clicked.connect(self.previousfile) #PREVIOUS
        self.ui.invertbutton.clicked.connect(self.invertdata) #INVERT

        ##########Event Navigation#########
        self.ui.previousbutton.clicked.connect(self.previousevent)
        self.ui.gobutton.clicked.connect(self.inspectevent)
        self.ui.nextbutton.clicked.connect(self.nextevent)

        ##########Axopatch Data############
        self.ui.ndChannel.clicked.connect(self.Plot)
        self.ui.plotBoth.clicked.connect(self.Plot)

        ##########IV Settings##############
        self.ui.IVxaxis.currentIndexChanged.connect(self.IVAxis)
        self.ui.IVyaxis.currentIndexChanged.connect(self.IVAxis)
        self.xaxisIV=self.ui.IVxaxis.currentIndex() #new variable for future work
        self.yaxisIV=self.ui.IVyaxis.currentIndex()

        ##########makeIV###################
        #self.ui.selectMakeIV.currentIndexChanged.connect(self.selectMakeIV)
        self.ui.makeIV.clicked.connect(self.makeIV)

        ##########pingpong###################
        self.ui.Prepingpong.clicked.connect(self.Prepingpong)
        self.ui.pingpong.clicked.connect(self.pingpong)
        ##########Pore Size -> Settings####  
        self.ui.saltbulkconductivity.valueChanged.connect(self.UpdateIV) # UpdateIV if changed. Works before any button pressed!
        self.ui.porelengthValue.valueChanged.connect(self.UpdateIV) # UpdateIV if changed

        ##########Pore Size -> Manual calculation####           
        self.ui.groupBox_5.clicked.connect(self.customCond)  #when manual calculation pressed
        self.ui.customCurrent.valueChanged.connect(self.UpdateIV) # UpdateIV if changed
        self.ui.customVoltage.valueChanged.connect(self.UpdateIV) # UpdateIV if changed
     
        ##########Main window########################
        ##########Figure saving -> save all##########
        self.ui.actionSave_All.triggered.connect(self.SaveAllFigures)####nothing happens

        ##########Display settings###################
        self.ui.actionUse_Clipping.triggered.connect(self.DisplaySettings)
        self.ui.actionUse_Downsampling.triggered.connect(self.DisplaySettings)
        self.ui.actionSave_IV_Data.triggered.connect(self.SaveIVData)

        ##########Event plot#########################
        self.ui.actionPlot_Common_Events.triggered.connect(self.EventFiltering)
        self.ui.actionPlot_i2_detected_only.triggered.connect(self.EventFiltering)
        self.ui.actionPlot_i1_detected_only.triggered.connect(self.EventFiltering)
        
        ###########Poresizer -> Find Pore size########
        self.ui.actionFind_pore_size.triggered.connect(self.sizethepore)  ####special func for finding the pore size see manual

        ###########Bathc Brocesser -> Batch it########
        self.ui.actionBatch_Process.triggered.connect(self.batchinfodialog)

        ###### Setting up plotting elements and their respective options######
        
        ###### All plots will have white background###
        self.ui.signalplot.setBackground('w')
        self.ui.voltageplotwin.setBackground('w')
        self.ui.eventplot.setBackground('w')
        self.ui.scatterplot.setBackground('w')
        self.ui.frachistplot.setBackground('w')
        self.ui.delihistplot.setBackground('w')
        self.ui.dwellhistplot.setBackground('w')
        self.ui.dthistplot.setBackground('w')
        self.ui.cutData.setBackground('w')
        self.ui.ivplot.setBackground('w')
        self.ui.powerSpecPlot.setBackground('w')
        self.ui.plot_pingpong.setBackground('w')
        #######If only one channel - channel 2 choice not available#####
        self.ui.AxopatchGroup.setVisible(0)

        ########working with Axis in Plotting at UsefulFunction.py######
        self.transverseAxis = pg.ViewBox()
        self.transverseAxisVoltage = pg.ViewBox()
        self.transverseAxisEvent = pg.ViewBox()

        ########Signal plot ##########################
        self.p1 = self.ui.signalplot #current vs time
        self.p1.enableAutoRange(axis = 'y')
        self.p1.enableAutoRange(axis='x')
        self.p1.setLabel('top', text = 'Signal plot')
        self.p1.setDownsampling(ds = True, auto = True, mode = 'subsample')
        self.p1.setClipToView(True)

        ########Voltage plot##########################
        self.voltagepl = self.ui.voltageplotwin
        self.voltagepl.enableAutoRange(axis = 'y')
        self.voltagepl.disableAutoRange(axis = 'x')
        self.voltagepl.setLabel('top', text = 'Voltage plot')
        self.voltagepl.setDownsampling(ds = True, auto = True, mode = 'subsample')
        self.voltagepl.setClipToView(True)
        self.voltagepl.setXLink(self.p1) #same links for signal plot and for voltageplot

        ########Event plot############################

        ###some feature added:
        ###When first launched, there is a Pyth-ion logo:
        self.p3 = self.ui.eventplot
        dir_path = os.path.dirname(os.path.realpath(__file__))
        self.logo = ndimage.imread(dir_path + os.sep + "pythionlogo.png")
        self.logo = np.rot90(self.logo,-1)
        self.logo = pg.ImageItem(self.logo)
        self.p3.addItem(self.logo)
        self.p3.setAspectLocked(True)
        self.p3.hideAxis('bottom')
        self.p3.hideAxis('left')
        ###will be replaced by Event plot, see Load() function for details

        ########Scatter plot##########################
        self.w1 = self.ui.scatterplot.addPlot()
        self.p2 = pg.ScatterPlotItem()
        self.p2.sigClicked.connect(self.clicked)
        self.w1.addItem(self.p2)
        self.w1.setLabel('bottom', text = 'Time', units = u'μs')
        self.w1.setLabel('left', text = 'Fractional Current Blockage')
        self.w1.setLogMode(x = True,y = False)
        self.w1.showGrid(x = True, y = True)
        self.cb = pg.ColorButton(self.ui.scatterplot, color = (0,0,255,50)) #small button left corner
        self.cb.setFixedHeight(30)
        self.cb.setFixedWidth(30)
        self.cb.move(0,250)
        self.cb.show()

        #########Frachist plot###########################
        self.w2 = self.ui.frachistplot.addPlot()  #frac plot
        self.w2.setLabel('bottom', text = 'Fractional Current Blockage')
        self.w2.setLabel('left', text = 'Counts')

        ##########Delihist plot##########################

        self.w3 = self.ui.delihistplot.addPlot() #deli histogram plot
        self.w3.setLabel('bottom', text = 'ΔI', units = 'A')
        self.w3.setLabel('left', text = 'Counts')

        ##########Dwell Hist Plot########################

        self.w4 = self.ui.dwellhistplot.addPlot() 
        self.w4.setLabel('bottom', text = 'Log Dwell Time', units = 'μs')
        self.w4.setLabel('left', text = 'Counts')

        ##########Dthistplot##############################

        self.w5 = self.ui.dthistplot.addPlot()
        self.w5.setLabel('bottom', text = 'dt', units = 's')
        self.w5.setLabel('left', text = 'Counts')

        ##########cut Data plot###########################

        self.cutplot = self.ui.cutData
        self.cutplot.setLabel('bottom', text = 'Time', units = 's')
        self.cutplot.setLabel('left', text = 'Voltage', units = 'V')
        self.cutplot.setLabel('top', text = 'Cut data plot')
        self.cutplot.enableAutoRange(axis = 'x')

        ##########IV plot##################################

        self.ivplota = self.ui.ivplot
        self.ivplota.setLabel('bottom', text = 'Current', units = 'A')
        self.ivplota.setLabel('left', text = 'Voltage', units = 'V')
        self.ivplota.enableAutoRange(axis = 'x')
        self.ivplota.setLabel('top', text = 'IV plot')

        ##########Power spectrum density plot##############
        self.psdplot = self.ui.powerSpecPlot
        self.psdplot.setLabel('left', 'PSD', units='(10^(-24))*A^2/Hz')
        self.psdplot.setLabel('bottom', 'Frequency', units='Hz')
        self.psdplot.setLabel('top', text='Power Spectrum Density')        

        ##########pingpong#########
        self.ui.previousbutton_pingpong.clicked.connect(self.previousevent_pingpong)
        self.ui.gobutton_pingpong.clicked.connect(self.inspectevent_pingpong)
        self.ui.nextbutton_pingpong.clicked.connect(self.nextevent_pingpong)
      
        ##########pingpong#########
        self.ui.load_and_merge_button_chimera.clicked.connect(self.load_and_merge_chimera) #loadbutton for merging chimera files


        ##########Setting up values for pore sizing#########
        
        self.useCustomConductance = 0
        self.conductance = 1e-9
        self.ui.resistanceText.setText('Resistance (1/G): ' + pg.siFormat(1/self.conductance, precision=5, suffix='Ohm', space=True, error=None, minVal=1e-25, allowUnicode=True))
        self.ui.conductanceText.setText('Conductance (G): ' + pg.siFormat(self.conductance, precision=5, suffix='S', space=True, error=None, minVal=1e-25, allowUnicode=True))
        self.ui.saltbulkconductivity.setOpts(value=10.5, suffix='S/m', siPrefix=True, decimals=6, step=1e-2)
        self.ui.porelengthValue.setOpts(value=0.7E-9, suffix='m', siPrefix=True, decimals=6, step=1e-11)
        self.ui.customCurrent.setOpts(value=10e-9, suffix='A', siPrefix=True, decimals=6, step=1e-10)
        self.ui.customVoltage.setOpts(value=500e-3, suffix='V', siPrefix=True, dec=True, step=10e-3, minStep=10e-3)
        self.ui.conductanceOutput.setText(pg.siFormat(10e-9/500e-3, precision=5, suffix='S', space=True, error=None, allowUnicode=True))
        self.ui.poresizeOutput.setText('Pore Size :' )

        #####Inserting pore size formula
        self.ui.PoreSizeformula.setBackground('w')
        self.PoreForm = self.ui.PoreSizeformula
        dir_path = os.path.dirname(os.path.realpath(__file__))
        self.formula = ndimage.imread(dir_path + os.sep + "poresizeformula.png")
        self.formula = np.rot90(self.formula,-1)
        self.formula = pg.ImageItem(self.formula)
        self.PoreForm.addItem(self.formula)
#        self.PoreForm.setAspectLocked(True)
        self.PoreForm.hideAxis('bottom')
        self.PoreForm.hideAxis('left')

        ##########Setting up coefficients for filtering and event decting for 1st channel######
        self.ui.LP_a.setOpts(value=0.999, suffix='', siPrefix=False, dec=True, step=1e-3, minStep=1e-4)
        self.ui.LP_S.setOpts(value=5, suffix='x STD', siPrefix=True, dec=True, step=10e-3, minStep=10e-3)
        self.ui.LP_E.setOpts(value=0, suffix='x STD', siPrefix=True, dec=True, step=10e-3, minStep=10e-3)
        self.ui.LP_eventlengthThresh.setOpts(value=1e-3, suffix='s', siPrefix=True, dec=True, step=10e-3, minStep=10e-3)

        ##########Setting up coefficients for filtering and event decting for 2nd channel######
        self.ui.LP_a_2.setOpts(value=0.999, suffix='', siPrefix=False, dec=True, step=1e-3, minStep=1e-4)
        self.ui.LP_S_2.setOpts(value=5, suffix='x STD', siPrefix=True, dec=True, step=10e-3, minStep=10e-3)
        self.ui.LP_E_2.setOpts(value=0, suffix='x STD', siPrefix=True, dec=True, step=10e-3, minStep=10e-3)
        self.ui.LP_eventlengthThresh_2.setOpts(value=1e-3, suffix='s', siPrefix=True, dec=True, step=10e-3, minStep=10e-3)

        self.Derivative = 'i2'

        ####### Initializing various variables used for analysis##############

        #######variables for downloading data
        self.direc = os.getcwd()
        self.datafilename = []

        #######Event analysis
        self.NumberOfEvents = 0
        self.UpwardsOn = 0
        self.AnalysisResults = {}
        self.sig = 'i1'

           
        #######Cut and baselinecalc    
        self.lr = [] #Linear region item
        self.hasbaselinebeenset = 0

        #######Stats graphics-> Scatter, Frac...######
        self.deli=[]
        self.frac=[]
        self.dwell=[]
        self.dt=[]
        self.catdata=[]
        self.colors=[]
        self.sdf = pd.DataFrame(columns = ['fn','color','deli','frac',
            'dwell','dt','startpoints','endpoints'])
    def loadfile(self): 
        self.ui.ndChannel.setChecked(False)
        self.ui.plotBoth.setChecked(False)
        Load.getfile(self)
    
    def UpdateIV(self): #Updating conductance and Pore Size data! Start before any button pressed
        
        if self.useCustomConductance:
            self.ui.conductanceOutput.setText(pg.siFormat(self.ui.customCurrent.value()/self.ui.customVoltage.value(), precision=5, suffix='S', space=True, error=None, allowUnicode=True))
            valuetoupdate=np.float(self.ui.customCurrent.value()/self.ui.customVoltage.value())
        else:
            valuetoupdate=self.conductance
        if self.ui.saltbulkconductivity.value():
            size=CalculatePoreSize(valuetoupdate, self.ui.porelengthValue.value(), self.ui.saltbulkconductivity.value())
            print('Size = '+ str(size))
            self.ui.poresizeOutput.setText('Pore Size: ' + pg.siFormat(size, precision=1, suffix='m', space=True, error=None,  allowUnicode=True))


    

    def Plot(self):
        # if self.hasbaselinebeenset==0:
        #     self.baseline=np.median(self.out['i1'])
        #     self.var=np.std(self.out['i1'])
        # self.ui.eventcounterlabel.setText('Baseline='+str(round(self.baseline*10**9, 2))+' nA')

        if not self.ui.plotBoth.isChecked():
            if self.ui.ndChannel.isChecked():
                self.sig = 'i2'
                self.sig2 = 'i1'
            else:
                self.sig = 'i1'
                self.sig2 = 'i2'
        
        if not self.ui.actionDon_t_Plot_if_slow.isChecked():
            if self.ui.plotBoth.isChecked():
                DoublePlot(self)
            else:
                PlotSingle(self)

    def IVAxis(self):
        self.xaxisIV = self.ui.IVxaxis.currentIndex()
        self.yaxisIV = self.ui.IVyaxis.currentIndex()




    def makeIV(self):
        self.selectMakeIV = self.ui.selectMakeIV.currentIndex()
        if self.yaxisIV == 0:
            xlab = 'i1'
        if self.yaxisIV == 1:
            xlab = 'i2'
        if self.xaxisIV == 0:
            ylab = 'v1'
        if self.xaxisIV == 1:
            ylab = 'v2'
        print('this works')
        if self.selectMakeIV == 0:
            print('makeIV linear')
        if self.selectMakeIV == 1:
            print('makeIV exp')
        self.cutplot.clear()
        (AllData, a) = uf.CutDataIntoVoltageSegments(self.out, extractedSegments = self.cutplot, x=xlab, y=ylab, delay=1, plotSegments=1)
        if AllData is not 0:
         #    Make IV
            (self.IVData, b) = uf.MakeIV(AllData, plot=1)
        
            #Fit IV
            self.ivplota.clear()
            (self.FitValues, iv) = uf.FitIV(self.IVData, x=xlab, y=ylab, iv=self.ivplota)
            self.conductance = self.FitValues['Slope']
            #self.UpdateIV()
            # Update Conductance

    def Prepingpong(self):
        self.Search_start = np.float(self.ui.Start_search_value.text()) #seconds from the beginning of the experiment
        self.Search_end = np.float(self.ui.End_search_value.text()) #seconds from the beginning of the experiment
        self.Search_start_points = int(self.Search_start * self.out['samplerate']) #Points from the beginning of the experiment
        self.Search_end_points = int(self.Search_end * self.out['samplerate']) #Points from the beginning of the experiment
       
        self.Delay_back_trace = np.float(self.ui.Delay_back_trace.text()) * 1e-3 #mSeconds for examinated reion selection
        self.Delay_back_trace_points = int(self.Delay_back_trace * self.out['samplerate']) #Points or examinated reion selection
        self.Safety_region = np.float(self.ui.Safety_region.text()) * 1e-3 #mSeconds for examinated reion selection
        self.Safety_region_points = int(self.Safety_region * self.out['samplerate']) #Points or examinated reion selection
        #Sigma_coeff = np.float(self.ui.sigma_coeff_value.text()) #coeff for algoritm
        
        self.zero_points = np.empty(0)  #points, where current passes through zero
        
        for j in np.arange(self.Search_start_points, self.Search_end_points): #checking for current passes through zero
            if np.sign(self.out['i1'][j-1] * self.out['i1'][j])< 0: 
                self.zero_points = np.append(self.zero_points, j)
        zero_points_time = self.zero_points/self.out['samplerate']
        self.p1.clear()
        self.p1.plot(self.t, self.out['i1'], pen='b')
        for n in np.arange(0,len(self.zero_points),1): #select regions for future CUSUM analysis
            self.control1 = int(self.zero_points[n] -self.Delay_back_trace_points)
            self.safety_reg = int(self.zero_points[n] -self.Delay_back_trace_points + self.Safety_region_points) 
            self.control2 =  int(self.zero_points[n])
            trace = self.out['i1'][self.control1:self.control2]
            self.p1.plot(self.t[self.control1:self.safety_reg], self.out['i1'][self.control1:self.safety_reg], pen='g')
            self.p1.plot(self.t[self.safety_reg:self.control2], self.out['i1'][self.safety_reg:self.control2], pen='r')

            
            ######eps for me
        #uf.PrintEps(self)
        print('Pre ping-pong worked fine')

        '''     
        l = len(zero_points_time) 
        IVData={}
        IVData['Mean'] = np.zeros(l)
        IVData['STD'] = np.zeros(l)

        table = OrderedDict([('Number', np.zeros(l)), ('Switch (s)', np.zeros(l)),('Start event (s)', np.zeros(l)), ('End event (s)', np.zeros(l)),('Duration (ms)', np.zeros(l)),('Depth (nA)', np.zeros(l))])


       

        for (n,k) in enumerate(zero_points):

            trace = self.out['i1'][int(zero_points[n] -2* Delay_back_points):int(zero_points[n])]
       
            IVData['Mean'][n] = np.mean(trace[0:int(0.5*Delay_back_points)])
            IVData['STD'][n] = np.std(trace[0:int(0.5*Delay_back_points)]) 
            delta = 1e-10
            h = delta / IVData['STD'][n]
            
            j=int(0.5*Delay_back_points)

            while np.abs(trace[j]- IVData['Mean'][n]) < Sigma_coeff * IVData['STD'][n] and j < (len(trace) -1): 
                IVData['Mean'][n] = np.mean(trace[0:int(j)])
                IVData['STD'][n] = np.std(trace[0:int(j)])        
                j+=1
                if j>  (len(trace) -2): print('Increase sigma coeff')

            table['Start event (s)'][n] = (zero_points[n] -2* Delay_back_points+j)/self.out['samplerate']
            start_event = j

            IVData['Mean'][n] = np.mean(trace[int(1.5*Delay_back_points):int(2*Delay_back_points)])
            IVData['STD'][n] = np.std(trace[int(1.5*Delay_back_points):int(2*Delay_back_points)])
            if n==9:
                print('Mean='+str(IVData['Mean'][n]))
                print('STD='+str(IVData['STD'][n]))
            j = int(1.5*Delay_back_points)
            while np.abs(trace[j]- IVData['Mean'][n]) < 4.5 * IVData['STD'][n]: 
                IVData['Mean'][n] = np.mean(trace[int(j):int(2*Delay_back_points)])
                IVData['STD'][n] = np.std(trace[int(j):int(2*Delay_back_points)])
                j-=1

            end_event = j

            #table['Number'][n] = n+1
            #table['End event (s)'][n] = (zero_points[n] -2* Delay_back_points+j)/self.out['samplerate']
            #table['Duration (ms)'][n] = (table['End event (s)'][n]-table['Start event (s)'][n])*1000
            #table['Depth (nA)'][n] = IVData['Mean'][n]*1e9  - np.mean(trace[start_event:end_event])*1e9 
            

        #table['Switch (s)'] = zero_points_time
        #file = open('results.txt','w') 
        #file.write(tabulate(table, headers='keys')) 
        #file.close() 
        ###########ANALYZE THE EVENT!!!!!##########
        #backPoint_coef = 15
        '''

    
     
    
    def pingpong(self):
        wb = Workbook()
        ws = wb.active
        wb.save("sample.xlsx")
        wb = load_workbook('sample.xlsx')
        ws = wb.get_active_sheet()
        
        ws.cell(row=1,column=1).value = 'Event number'
        ws.cell(row=1,column=2).value = 'Switch (s)'
        ws.cell(row=1,column=3).value = 'Start event (s)'
        ws.cell(row=1,column=4).value = 'End event (s)'
        ws.cell(row=1,column=5).value = 'Duration (ms)'
        ws.cell(row=1,column=6).value = 'Depth (nA)'
        ws.cell(row=1,column=7).value = 'Switch (in front) - End event (ms)'
        ws.cell(row=1,column=8).value = 'Start event - Switch back (ms)'
        
        for m in np.arange(2,100,1):
            for k in np.arange(1,100,1):  
                ws.cell(row=m,column=k).value = ''
            
        for n in np.arange(0,len(self.zero_points),1): 
            self.control1 = int(self.zero_points[n] -self.Delay_back_trace_points)
            self.safety_reg = int(self.zero_points[n] -self.Delay_back_trace_points + self.Safety_region_points) 
            self.control2 =  int(self.zero_points[n])
            trace = self.out['i1'][self.control1:self.control2]
            self.var=np.std(trace[:self.safety_reg])
            self.data = trace
            cusum_results = self.CUSUM()
            Event = self.control1  +cusum_results['EventDelay']
            
            if len(Event)>2:
                #####filling exel file with numbers
                ws.cell(row=n+2,column=1).value = n+1 #event number 
                ws.cell(row=n+2,column=2).value = "%6.4f"% float(self.zero_points[n] / self.out['samplerate']) #zero points
                ws.cell(row=n+2,column=3).value = "%6.4f"% float(Event[1] / self.out['samplerate']) #Start event    
                ws.cell(row=n+2,column=4).value = "%6.4f"% float(Event[-2] / self.out['samplerate']) #End event
                ws.cell(row=n+2,column=5).value = "%3.2f"% float((Event[-2] - Event[1]) *1e3 / self.out['samplerate']) #Event duration
                ws.cell(row=n+2,column=6).value = "%3.1f"% float((cusum_results['CurrentLevels'][1] - cusum_results['CurrentLevels'][0]) * 1e9) #Depth
                ws.cell(row=n+2,column=7).value = "%3.1f"% float((self.zero_points[n] - Event[-2]) * 1e3 /self.out['samplerate'] ) #Switch (in front) - End event (ms)
                ws.cell(row=n+2,column=8).value = "%6.2f"%  float((Event[1] - self.zero_points[n-1]) * 1e3 /self.out['samplerate'] ) #Start event - Switch back (ms)
                for interm_event in np.arange(2,len(cusum_results['EventDelay'])-2):
                    ws.cell(row=1,column=8+interm_event).value = 'Intermediate events'
                    ws.cell(row=2+n,column=8+interm_event).value = "%6.4f"% float(Event[interm_event]/ self.out['samplerate']) #intermediate events



            #ws.cell(row=n+2,column=3+len(cusum_results['EventDelay'])).value = (cusum_results['EventDelay'][-1]-cusum_results['EventDelay'][0])/ self.out['samplerate']
            
            
            self.p1.plot(self.t[self.control1:Event[1]], np.repeat(np.array([cusum_results['CurrentLevels'][0]]), len(self.t[self.control1:Event[1]])), pen=pg.mkPen(color=(173, 27, 183), width=3))
            for number_cusum in np.arange(1,len(Event)-1 ,1): 
                self.p1.plot(self.t[Event[number_cusum]:Event[number_cusum+1]], np.repeat(np.array([cusum_results['CurrentLevels'][number_cusum]]), len(self.t[Event[number_cusum]:Event[number_cusum+1]])), pen=pg.mkPen(color=(173, 27, 183), width=3))
                self.p1.plot(np.linspace(self.t[Event[number_cusum]],self.t[Event[number_cusum]] + 0.00001 ,100),np.linspace(cusum_results['CurrentLevels'][number_cusum-1],cusum_results['CurrentLevels'][number_cusum],100), pen=pg.mkPen(color=(173, 27, 183), width=3))
            self.p1.plot(self.t[Event[-2]:self.control2], np.repeat(np.array([cusum_results['CurrentLevels'][-1]]), len(self.t[Event[-2]:self.control2])), pen=pg.mkPen(color=(173, 27, 183), width=3))
            self.p1.autoRange()
            
        wb.save("sample.xlsx")        
            #####write to Exel########
            #wb = Workbook() # create exel File
            

            #self.p1.plot(self.t[self.control1:self.safety_reg], self.out['i1'][self.control1:self.safety_reg], pen='g')
            #self.p1.plot(self.t[self.safety_reg:self.control2], self.out['i1'][self.safety_reg:self.control2], pen='r')
        #self.ui.plot_pingpong.plot(np.linspace(self.safety_reg ,self.safety_reg + 0.00001, 400),np.linspace(trace[0]-3e-9,trace[0]+3e-9,400), color='red', width = 40)
        #self.ui.plot_pingpong.plot(self.t[self.control1:self.control2], trace , pen='b')
        
        #self.var=np.std(trace[:self.safety_reg])
        #self.ui.plot_pingpong.plot(self.t[self.control1:self.control2], self.out['i1'][self.control1:self.control2] , pen='b')
        #self.data = trace
        #cusum_results = self.CUSUM()
        #Event = self.control1  +cusum_results['EventDelay']      
        print('Ping-pong worked fine')
      

    def Analyze(self):
        #self.p1.clear()
        #self.p1.setDownsampling(ds = False)
        min_duration_points = int(np.float64(self.ui.min_duration.text()) * 1e-3 * self.outputsamplerate)
        self.cusum = {}
        self.cusum['i1'] = detection(self, self.data['i1'], dt = 1/self.outputsamplerate, threshhold  = np.float64(self.ui.levelthresholdentry.text()) * 1e-9, minlength = min_duration_points, maxstates = 10000)
        try: 
            self.cusum['i2'] = detection(self, self.data['i2'], dt = 1/self.outputsamplerate, threshhold  = np.float64(self.ui.levelthresholdentry.text()) * 1e-9, minlength = min_duration_points, maxstates = 10000)
        except KeyError: 
            pass
        print_fitting(self, self.cusum['i1'])
        print('everythink ok')
        
        return 0

        
        
    def CUSUM(self):
        #self.p1.clear()
        #self.p1.setDownsampling(ds = False)
        min_duration_points = int(np.float64(self.ui.min_duration.text()) * 1e-3 * self.outputsamplerate)
        cusum = detect_cusum(self, self.data,  dt = 1/self.outputsamplerate, threshhold  = np.float64(self.ui.levelthresholdentry.text()) * 1e-9, minlength = min_duration_points, maxstates = 10)
        print('everythink ok')
        return cusum

    def inspectevent_pingpong(self):    ######Individual inspection of the event

        
        n = np.int(self.ui.pingpong_event_number.text()) -1
        '''
        ######printing event info##########print("%10.7f"% float(startpoints[j]/self.out['samplerate']))
        self.ui.starteventtext.setText('Start event (s): ' + str("%6.4f"% float(self.table_info['Start event (s)'][eventnumber])))
        self.ui.endeventtext.setText('END event(s): ' + str("%6.4f"% float(self.table_info['End event (s)'][eventnumber])))
        self.ui.switchtext.setText('Switch (s): ' + str("%6.4f"% float(self.table_info['Switch (s)'][eventnumber])))
        self.ui.switchend.setText('Switch (in front)-End event (ms): ' + str("%3.1f"% float( (self.table_info['Switch (s)'][eventnumber]-self.table_info['End event (s)'][eventnumber]) * 1e3 )))
        self.ui.startswitch.setText('Start event - Switch (back) (ms): ' + str("%3.0f"% float((self.table_info['Start event (s)'][eventnumber]-self.table_info['Switch (s)'][eventnumber-1]) * 1e3) ))
        self.ui.durationtext.setText('Duration (ms): ' + str("%3.2f"% float((self.table_info['End event (s)'][eventnumber]-self.table_info['Start event (s)'][eventnumber]) * 1e3) ))
        
        self.ui.depthtext.setText('Depth (nA):: ' + str("%4.2f"% float(self.table_info['Depth (nA)'][eventnumber])))
        eventbuffer = 150
        self.ui.plot_pingpong.setLabel('bottom', text='Time', units='s')
        self.ui.plot_pingpong.setLabel('left', text='Current', units='A')
        self.ui.plot_pingpong.clear()
        self.ui.plot_pingpong.plot(self.t[int(self.startpoints[eventnumber] - eventbuffer):int(self.endpoints[eventnumber] + eventbuffer)], self.out['i1'][int(self.startpoints[eventnumber] - eventbuffer):int(self.endpoints[eventnumber] + eventbuffer)] , pen='b')

        self.ui.plot_pingpong.plot(self.t[int(self.startpoints[eventnumber] - eventbuffer):int(self.endpoints[eventnumber] + eventbuffer)], np.concatenate((np.repeat(np.array([self.localBaseline[eventnumber]]), eventbuffer), np.repeat(np.array([self.localBaseline[eventnumber] - self.depth[eventnumber]]), self.endpoints[eventnumber] - self.startpoints[eventnumber]), np.repeat(np.array([self.localBaseline[eventnumber]]), eventbuffer)), 0), pen=pg.mkPen(color=(173, 27, 183), width=3))
        self.ui.plot_pingpong.autoRange()
        '''
        self.ui.plot_pingpong.clear()
        Delay_back = np.float(0.1)  #Seconds how far event should be from switch
        Delay_back_points = int(Delay_back * self.out['samplerate']) #Points how far event should be from switch
        self.control1 = int(self.zero_points[n] -Delay_back_points)
        self.control2 =  int(self.zero_points[n])
        trace = self.out['i1'][self.control1:self.control2]
        self.var=np.std(trace[:1000])
        self.ui.plot_pingpong.plot(self.t[self.control1:self.control2], self.out['i1'][self.control1:self.control2] , pen='b')
        self.data = trace
        cusum_results = self.CUSUM()
        mask = np.ones(len(cusum_results['EventDelay']), dtype = bool)
        mask[[0,-1]]= False #cut off 0 and last elements
        Event = self.control1  +cusum_results['EventDelay']
        #print('Masked='+str(Event))
        #Level = cusum_results['CurrentLevels'][0]
        #self.ui.plot_pingpong.plot(np.linspace(Event1,Event1 + 0.00001 ,100),np.linspace(Level-3e-9,Level+3e-9,100), color='green')
        #self.ui.plot_pingpong.plot(np.linspace(Event2,Event2 + 0.00001 ,100),np.linspace(Level-3e-9,Level+3e-9,100), color='green')
        #self.ui.plot_pingpong.plot(np.linspace(Event3,Event3 + 0.00001 ,100),np.linspace(Level-3e-9,Level+3e-9,100), color='green')

        #self.ui.plot_pingpong.clear()
 

        self.ui.plot_pingpong.plot(self.t[self.control1:Event[1]], np.repeat(np.array([cusum_results['CurrentLevels'][0]]), len(self.t[self.control1:Event[1]])), pen=pg.mkPen(color=(173, 27, 183), width=3))
        for number_cusum in np.arange(1,len(Event)-1 ,1): 
            self.ui.plot_pingpong.plot(self.t[Event[number_cusum]:Event[number_cusum+1]], np.repeat(np.array([cusum_results['CurrentLevels'][number_cusum]]), len(self.t[Event[number_cusum]:Event[number_cusum+1]])), pen=pg.mkPen(color=(173, 27, 183), width=3))
            self.ui.plot_pingpong.plot(np.linspace(self.t[Event[number_cusum]],self.t[Event[number_cusum]] + 0.00001 ,100),np.linspace(cusum_results['CurrentLevels'][number_cusum-1],cusum_results['CurrentLevels'][number_cusum],100), pen=pg.mkPen(color=(173, 27, 183), width=3))
        self.ui.plot_pingpong.plot(self.t[Event[-2]:self.control2], np.repeat(np.array([cusum_results['CurrentLevels'][-1]]), len(self.t[Event[-2]:self.control2])), pen=pg.mkPen(color=(173, 27, 183), width=3))
        self.ui.plot_pingpong.autoRange()

    def previousevent_pingpong(self):
        self.ui.pingpong_event_number.setText(str(np.int(self.ui.pingpong_event_number.text())-1))
        if np.int(self.ui.pingpong_event_number.text()) < 1: self.ui.pingpong_event_number.setText('1')
        self.inspectevent_pingpong()
    def nextevent_pingpong(self):
        self.ui.pingpong_event_number.setText(str(np.int(self.ui.pingpong_event_number.text())+1))
        self.inspectevent_pingpong()

    def baselinecalc(self):
        self.threshold  = np.float64(self.ui.levelthresholdentry.text()) * 1e-9
        if self.lr==[]:
            self.p1.clear()
            self.lr = pg.LinearRegionItem()
            self.lr.hide()
            self.p1.addItem(self.lr)

#            self.p1.plot(self.t[::100],self.data[::100],pen='b')
            self.p1.plot(self.t, self.data, pen='b')
            self.lr.show()

        else:
            calcregion=self.lr.getRegion()
            self.p1.clear()

            self.baseline=np.median(self.data[np.arange(np.int(calcregion[0]*self.outputsamplerate),np.int(calcregion[1]*self.outputsamplerate))])
            self.var=np.std(self.data[np.arange(np.int(calcregion[0]*self.outputsamplerate),np.int(calcregion[1]*self.outputsamplerate))])
#            self.p1.plot(self.t[::10][2:][:-2],self.data[::10][2:][:-2],pen='b')
            self.p1.plot(self.t,self.data,pen='b')
            self.p1.addLine(y=self.baseline,pen='g')
            self.p1.addLine(y=self.baseline - self.threshold,pen='r')
            self.p1.addLine(y=self.baseline + self.threshold,pen='r')
            self.lr=[] #clear linear region
            self.hasbaselinebeenset=1 #mark as baseline set
            self.ui.eventcounterlabel.setText('Baseline='+str(round(self.baseline*10**9,2))+' nA')
            self.p1.autoRange()


    def CutData(self):
        cut(self)

    def SaveIVData(self):
        uf.ExportIVData(self)

    def analyze(self):
        self.coefficients = {}
        self.coefficients['i1'] = {'a': np.float(self.ui.LP_a.value()), 'E': np.float(self.ui.LP_E.value()),
                                   'S': np.float(self.ui.LP_S.value()),
                                   'eventlengthLimit': np.float(self.ui.LP_eventlengthThresh.value()) * self.outputsamplerate}
        if self.data['graphene']:
            self.coefficients['i2'] = {'a': np.float(self.ui.LP_a_2.value()), 'E': np.float(self.ui.LP_E_2.value()),
                                 'S': np.float(self.ui.LP_S_2.value()),
                                 'eventlengthLimit': np.float(self.ui.LP_eventlengthThresh_2.value()) * self.outputsamplerate}
            chan = ['i1', 'i2']
            start1 = timer()
            for sig in chan:
                self.AnalysisResults[sig] = {}
                self.AnalysisResults[sig]['RoughEventLocations'] = uf.RecursiveLowPassFast(self.data[sig], self.coefficients[sig], self)
                if self.UpwardsOn:
                    self.AnalysisResultsUp[sig+'_Up'] = {}
                    self.AnalysisResultsUp[sig+'_Up']['RoughEventLocations'] = uf.RecursiveLowPassFastUp(self.data[self.sig], self.coefficients[sig])


            end1 = timer()
            print('The Low-pass took {} s on both channels.'.format(str(end1 - start1)))
            self.sig = 'i1'
            uf.AddInfoAfterRecursive(self)
            self.sig = 'i2'
            uf.AddInfoAfterRecursive(self)
            end2 = timer()
            print('Adding Info took {} s on both channels.'.format(str(end2-end1)))

            uf.SavingAndPlottingAfterRecursive(self)
            uf.SaveToHDF5(self)
            end3 = timer()
            print('Saving took {} s on both channels.'.format(str(end3-end2)))

            if 'i1' in self.AnalysisResults and 'i2' in self.AnalysisResults:
                (self.CommonIndexes, self.OnlyIndexes) = uf.CombineTheTwoChannels(self.matfilename + '_OriginalDB.hdf5')
                print('Two channels are combined')
                uf.EditInfoText(self)
                self.EventFiltering(self)
            end4 = timer()
            print('Combining took {} s on both channels.'.format(str(end4-end3)))
        else:
            start1 = timer()
            self.AnalysisResults['i1'] = {}
            self.AnalysisResults['i1']['RoughEventLocations'] = uf.RecursiveLowPassFast(self.data['i1'], self.coefficients['i1'], self)
            if self.UpwardsOn:
                self.AnalysisResults['i1_Up']['RoughEventLocations'] = uf.RecursiveLowPassFast(self.data['i1'],
                                                                                            self.coefficients['i1'])
            end1 = timer()
            print('The Low-pass took {} s on both channels.'.format(str(start1 - end1)))
            self.sig = 'i1'
            uf.AddInfoAfterRecursive(self)
            end2 = timer()
            print('Adding Info took {} s on both channels.'.format(str(end2 - end1)))
            self.NumberOfEvents=len(np.uint64(self.AnalysisResults['i1']['RoughEventLocations'][:, 0]))
            uf.SaveToHDF5(self)
            end3 = timer()
            print('Saving took {} s on both channels.'.format(str(end3 - end2)))

    def inspectevent(self, clicked = []):
        if self.data['graphene']:
            if self.ui.actionPlot_Common_Events.isChecked():
                uf.PlotEventDoubleFit(self, clicked)
            elif self.ui.actionPlot_i1_detected_only.isChecked() or self.ui.actionPlot_i2_detected_only.isChecked():
                uf.PlotEventDouble(self)
        else:
            #uf.PlotEventSingle_CUSUM(self, clicked)
            uf.PlotEventSingle(self, clicked)

    def nextevent(self):
        eventnumber=np.int(self.ui.eventnumberentry.text())
        print(self.NumberOfEvents)
        if eventnumber>=self.NumberOfEvents-1:
            eventnumber=0
        else:
            eventnumber=np.int(self.ui.eventnumberentry.text())+1
        self.ui.eventnumberentry.setText(str(eventnumber))
        self.inspectevent()

    def previousevent(self):
        eventnumber=np.int(self.ui.eventnumberentry.text())
        print(self.NumberOfEvents)
        if eventnumber<=0:
            eventnumber=self.NumberOfEvents-1
        else:
            eventnumber=np.int(self.ui.eventnumberentry.text())-1
        self.ui.eventnumberentry.setText(str(eventnumber))
        self.inspectevent()

   

    def load_and_merge_chimera(self):
        load_and_merge(self)

        '''print('load and merge')
        self.ui.Text_Chimera.setText('Filename: ')
        try:
            ######## attempt to open dialog from most recent directory########
            datafilenametemp = QtGui.QFileDialog.getOpenFileName(parent=self, caption='Open file', directory=str(self.direc), filter="Amplifier Files(*.dat *.log  *.mat)") #creates turple with two values: the adress of the selected file and the line of Amplifier Filters
            if datafilenametemp[0] != '': #if you select some file the param will be not 0
                self.datafilename=datafilenametemp[0] #full path to the loaded file
                self.direc=os.path.dirname(self.datafilename) #Return the directory name
                self.Load() #Load module
        except IOError:
            #### if user cancels during file selection, exit loop#############
            pass'''

    def clearscatter(self):
        self.p2.setData(x=[],y=[])
        self.lastevent=[]
        self.ui.scatterplot.update()
        self.w2.clear() #clear frac plot
        self.w3.clear() #deli histogram plot
        self.w4.clear() #Dwell histogram plot
        self.w5.clear() #Dt histogram plot
        self.sdf = pd.DataFrame(columns = ['fn','color','deli','frac',
            'dwell','dt','startpoints','endpoints'])

    def deleteevent(self):
        eventnumber = np.int(self.ui.eventnumberentry.text())
        firstindex = self.sdf.fn[self.sdf.fn == self.matfilename].index[0]
        if eventnumber > self.numberofevents:
            eventnumber = self.numberofevents-1
            self.ui.eventnumberentry.setText(str(eventnumber))
        self.deli=np.delete(self.deli,eventnumber)
        self.dwell=np.delete(self.dwell,eventnumber)
        self.dt=np.delete(self.dt,eventnumber)
        self.frac=np.delete(self.frac,eventnumber)
        self.startpoints=np.delete(self.startpoints, eventnumber)
        self.endpoints=np.delete(self.endpoints, eventnumber)
        self.p2.data=np.delete(self.p2.data,firstindex + eventnumber)

        self.numberofevents = len(self.dt)
        self.ui.eventcounterlabel.setText('Events:'+str(self.numberofevents))

        self.sdf = self.sdf.drop(firstindex + eventnumber).reset_index(drop = True)
        self.inspectevent()

        self.w2.clear()
        self.w3.clear()
        self.w4.clear()
        self.w5.clear()
        colors = self.sdf.color
        for i, x in enumerate(colors):
            fracy, fracx = np.histogram(self.sdf.frac[self.sdf.color == x], bins=np.linspace(0, 1, int(self.ui.fracbins.text())))
            deliy, delix = np.histogram(self.sdf.deli[self.sdf.color == x], bins=np.linspace(float(self.ui.delirange0.text())*10**-9, float(self.ui.delirange1.text())*10**-9, int(self.ui.delibins.text())))
            dwelly, dwellx = np.histogram(np.log10(self.sdf.dwell[self.sdf.color == x]), bins=np.linspace(float(self.ui.dwellrange0.text()), float(self.ui.dwellrange1.text()), int(self.ui.dwellbins.text())))
            dty, dtx = np.histogram(self.sdf.dt[self.sdf.color == x], bins=np.linspace(float(self.ui.dtrange0.text()), float(self.ui.dtrange1.text()), int(self.ui.dtbins.text())))

#            hist = pg.PlotCurveItem(fracy, fracx , stepMode = True, fillLevel=0, brush = x, pen = 'k')
#            self.w2.addItem(hist)

            hist = pg.BarGraphItem(height = fracy, x0 = fracx[:-1], x1 = fracx[1:], brush = x)
            self.w2.addItem(hist)

#            hist = pg.PlotCurveItem(delix, deliy , stepMode = True, fillLevel=0, brush = x, pen = 'k')
#            self.w3.addItem(hist)

            hist = pg.BarGraphItem(height = deliy, x0 = delix[:-1], x1 = delix[1:], brush = x)
            self.w3.addItem(hist)
#            self.w3.autoRange()
            self.w3.setRange(xRange = [float(self.ui.delirange0.text())*10**-9, float(self.ui.delirange1.text())*10**-9])

#            hist = pg.PlotCurveItem(dwellx, dwelly , stepMode = True, fillLevel=0, brush = x, pen = 'k')
#            self.w4.addItem(hist)

            hist = pg.BarGraphItem(height = dwelly, x0 = dwellx[:-1], x1 = dwellx[1:], brush = x)
            self.w4.addItem(hist)

#            hist = pg.PlotCurveItem(dtx, dty , stepMode = True, fillLevel=0, brush = x, pen = 'k')
#            self.w5.addItem(hist)

            hist = pg.BarGraphItem(height = dty, x0 = dtx[:-1], x1 = dtx[1:], brush = x)
            self.w5.addItem(hist)

        self.save()
        uf.SaveToHDF5(self)

    def invertdata(self):
        self.p1.clear()
        self.data['i1'] = - self.data['i1']
        self.var['i1'] = np.std(self.data['i1'])
        try:
            self.data['i2']= - self.data['i2']
            self.data['v2'] = - self.data['v2']
            self.var['i2'] = np.std(self.data['i2'])
        except KeyError: 
            pass       
        self.Plot()

    def clicked(self, plot, points):
        for i, p in enumerate(self.p2.points()):
            if p.pos() == points[0].pos():
                clickedindex = i

        if self.sdf.fn[clickedindex] != self.matfilename:
            print('Event is from an earlier file, not clickable')

        else:
            self.inspectevent(clickedindex)

    def concatenatetext(self):
        if self.direc==[]:
            textfilenames = QtGui.QFileDialog.getOpenFileNames(self, 'Open file','*.txt')
            self.direc=os.path.dirname(str(textfilenames[0]))
        else:
            textfilenames =QtGui.QFileDialog.getOpenFileNames(self, 'Open file',self.direc,'*.txt')
            self.direc=os.path.dirname(str(textfilenames[0]))
        i=0
        while i<len(textfilenames):
            temptextdata=np.fromfile(str(textfilenames[i]),sep='\t')
            temptextdata=np.reshape(temptextdata,(len(temptextdata)/4,4))
            if i==0:
                newtextdata=temptextdata
            else:
                newtextdata=np.concatenate((newtextdata,temptextdata))
            i=i+1

        newfilename = QtGui.QFileDialog.getSaveFileName(self, 'New File name',self.direc,'*.txt')
        np.savetxt(str(newfilename),newtextdata,delimiter='\t')

    def nextfile(self):
        if str(os.path.splitext(self.datafilename)[1])=='.log':
            startindex=self.matfilename[-6::]
            filebase=self.matfilename[0:len(self.matfilename)-6]
            nextindex=str(int(startindex)+1)
            while os.path.isfile(filebase+nextindex+'.log')==False:
                nextindex=str(int(nextindex)+1)
                if int(nextindex)>int(startindex)+1000:
                    print('no such file')
                    break
            if os.path.isfile(filebase+nextindex+'.log')==True:
                self.datafilename=(filebase+nextindex+'.log')
                self.Load()

        if str(os.path.splitext(self.datafilename)[1])=='.abf':
            startindex=self.matfilename[-4::]
            filebase=self.matfilename[0:len(self.matfilename)-4]
            nextindex=str(int(startindex)+1).zfill(4)
            while os.path.isfile(filebase+nextindex+'.abf')==False:
                nextindex=str(int(nextindex)+1).zfill(4)
                if int(nextindex)>int(startindex)+1000:
                    print('no such file')
                    break
            if os.path.isfile(filebase+nextindex+'.abf')==True:
                self.datafilename=(filebase+nextindex+'.abf')
                self.Load()

    def previousfile(self):
        if str(os.path.splitext(self.datafilename)[1])=='.log':
            startindex=self.matfilename[-6::]
            filebase=self.matfilename[0:len(self.matfilename)-6]
            nextindex=str(int(startindex)-1)
            while os.path.isfile(filebase+nextindex+'.log')==False:
                nextindex=str(int(nextindex)-1)
                if int(nextindex)<int(startindex)-1000:
                    print('no such file')
                    break
            if os.path.isfile(filebase+nextindex+'.log')==True:
                self.datafilename=(filebase+nextindex+'.log')
                self.Load()

        if str(os.path.splitext(self.datafilename)[1])=='.abf':
            startindex=self.matfilename[-4::]
            filebase=self.matfilename[0:len(self.matfilename)-4]
            nextindex=str(int(startindex)-1).zfill(4)
            while os.path.isfile(filebase+nextindex+'.abf')==False:
                nextindex=str(int(nextindex)-1).zfill(4)
                if int(nextindex)<int(startindex)-1000:
                    print('no such file')
                    break
            if os.path.isfile(filebase+nextindex+'.abf')==True:
                self.datafilename=(filebase+nextindex+'.abf')
                self.Load()

    def savetrace(self):
        self.data.astype('d').tofile(self.matfilename+'_trace.bin')

    def showcattrace(self):
        eventbuffer=np.int(self.ui.eventbufferentry.value())
        numberofevents=len(self.dt)

        self.p1.clear()
        eventtime = [0]
        for i in range(numberofevents):
            if i<numberofevents-1:
                if endpoints[i]+eventbuffer>startpoints[i+1]:
                    print('overlapping event')
                else:
                    eventdata = self.data[startpoints[i]-eventbuffer:endpoints[i]+eventbuffer]
                    fitdata = np.concatenate((np.repeat(np.array([self.localBaseline[i]]),eventbuffer),np.repeat(np.array([self.localBaseline[i]-self.deli[i]]),endpoints[i]-startpoints[i]),np.repeat(np.array([self.localBaseline[i]]),eventbuffer)),0)
                    eventtime = np.arange(0,len(eventdata)) + .75*eventbuffer + eventtime[-1]
                    self.p1.plot(eventtime/self.outputsamplerate, eventdata,pen='b')
                    self.p1.plot(eventtime/self.outputsamplerate, fitdata,pen=pg.mkPen(color=(173,27,183),width=2))

        self.p1.autoRange()

    def savecattrace(self):
        eventbuffer=np.int(self.ui.eventbufferentry.value())
        numberofevents=len(self.dt)
        self.catdata=self.data[startpoints[0]-eventbuffer:endpoints[0]+eventbuffer]
        self.catfits=np.concatenate((np.repeat(np.array([self.baseline]),eventbuffer),np.repeat(np.array([
            self.baseline-self.deli[0]]),endpoints[0]-startpoints[0]),
            np.repeat(np.array([self.baseline]),eventbuffer)),0)

        for i in range(numberofevents):
            if i<numberofevents-1:
                if endpoints[i]+eventbuffer>startpoints[i+1]:
                    print('overlapping event')
                else:
                    self.catdata=np.concatenate((self.catdata,self.data[startpoints[i]-eventbuffer:endpoints[i]+eventbuffer]),0)
                    self.catfits=np.concatenate((self.catfits,np.concatenate((np.repeat(np.array([self.baseline]),eventbuffer),np.repeat(np.array([
                        self.baseline-self.deli[i]]),endpoints[i]-startpoints[i]),np.repeat(np.array([self.baseline]),eventbuffer)),0)),0)

        self.tcat=np.arange(0,len(self.catdata))
        self.tcat=self.tcat/self.outputsamplerate
        self.catdata=self.catdata[::10]
        self.catdata.astype('d').tofile(self.matfilename+'_cattrace.bin')

    def keyPressEvent(self, event):
        key = event.key()
        if key == QtCore.Qt.Key_Up:
            self.nextfile()
        if key == QtCore.Qt.Key_Down:
            self.previousfile()
        if key == QtCore.Qt.Key_Right:
            self.nextevent()
        if key == QtCore.Qt.Key_Left:
            self.previousevent()
    #    if key == QtCore.Qt.Key_Return:
    #        self.Load()
        if key == QtCore.Qt.Key_Space:
            self.analyze()
        if key == QtCore.Qt.Key_Delete:
            self.deleteevent()
        if key == QtCore.Qt.Key_S:
            self.skeypressed()
        if key == QtCore.Qt.Key_D:
            self.dkeypresses()

    def saveeventfits(self):
        eventbuffer=np.int(self.ui.eventbufferentry.value())
        numberofevents=len(self.dt)
        self.catdata=self.data[startpoints[0]-eventbuffer:endpoints[0]+eventbuffer]
        self.catfits=np.concatenate((np.repeat(np.array([self.baseline]),eventbuffer),np.repeat(np.array([
            self.baseline-self.deli[0]]),endpoints[0]-startpoints[0]),
            np.repeat(np.array([self.baseline]),eventbuffer)),0)

        for i in range(numberofevents):
            if i<numberofevents-1:
                if endpoints[i]+eventbuffer>startpoints[i+1]:
                    print('overlapping event')
                else:
                    self.catdata=np.concatenate((self.catdata,self.data[startpoints[i]-eventbuffer:endpoints[i]+eventbuffer]),0)
                    self.catfits=np.concatenate((self.catfits,np.concatenate((np.repeat(np.array([self.baseline]),eventbuffer),np.repeat(np.array([
                        self.baseline-self.deli[i]]),endpoints[i]-startpoints[i]),np.repeat(np.array([self.baseline]),eventbuffer)),0)),0)

        self.tcat=np.arange(0,len(self.catdata))
        self.tcat=self.tcat/self.outputsamplerate
        self.catfits.astype('d').tofile(self.matfilename+'_cattrace.bin')

   
    def savetarget(self):
        self.batchinfo = self.batchinfo.append(pd.DataFrame({'deli':self.deli,
                    'frac':self.frac,'dwell':self.dwell,'dt':self.dt, 
                    'startpoints':startpoints,'endpoints':endpoints}), ignore_index=True)
        self.batchinfo.to_pickle(self.matfilename+'batchinfo.pkl')

    def batchinfodialog(self):
        self.bp = batchprocesser()
        self.bp.show()
        
        QtCore.QObject.connect(self.bp.uibp.okbutton, QtCore.SIGNAL('clicked()'), self.batchprocess)
        
    def batchprocess(self):
        global endpoints, startpoints
        
        self.p1.setDownsampling(ds = False)
        self.mindwell = np.float64(self.bp.uibp.mindwellbox.text())
        self.minfrac = np.float64(self.bp.uibp.minfracbox.text())
        self.minlevelt = np.float64(self.bp.uibp.minleveltbox.text())*10**-6
        self.samplerate = self.bp.uibp.sampratebox.text()
        self.LPfiltercutoff = self.bp.uibp.LPfilterbox.text()
        self.ui.outputsamplerateentry.setText(self.samplerate)
        self.ui.LPentry.setText(self.LPfiltercutoff)
        cusumstep = np.float64(self.bp.uibp.cusumstepentry.text())
        cusumthresh = np.float64(self.bp.uibp.cusumthreshentry.text())
        self.bp.destroy()   
        self.p1.clear()
        
        try:
            ######## attempt to open dialog from most recent directory########
            self.filelist = QtGui.QFileDialog.getOpenFileNames(self,'Select Files',self.direc,("*.pkl"))
            self.direc=os.path.dirname(self.filelist[0])
        except TypeError:
            ####### if no recent directory exists open from working directory##
            self.direc==[]
            self.filelist = QtGui.QFileDialog.getOpenFileNames(self, 'Select Files',os.getcwd(),("*.pkl"))
            self.direc=os.path.dirname(self.filelist[0])
        except IOError:
            #### if user cancels during file selection, exit loop#############
            return

        eventbuffer=np.int(self.ui.eventbufferentry.vlaue())
        eventtime = [0]
        ll = np.array([])


        for f in self.filelist: 
            batchinfo = pd.read_pickle(f)
            try:
                self.datafilename = f[:-13] + '.opt'
                self.Load(loadandplot = False)
            except IOError:
                self.datafilename = f[:-13] + '.log'
                self.Load(loadandplot = False)
                
            
            try:
                cs = batchinfo.cutstart[np.isfinite(batchinfo.cutstart)]
                ce = batchinfo.cutend[np.isfinite(batchinfo.cutend)]
                for i, cut in enumerate(cs):
                    self.data = np.delete(self.data,np.arange(np.int(cut*self.outputsamplerate),np.int(ce[i]*self.outputsamplerate)))
            except TypeError:
                pass
             
             
            self.deli = np.array(batchinfo.deli[np.isfinite(batchinfo.deli)])
            self.frac = np.array(batchinfo.frac[np.isfinite(batchinfo.frac)])
            self.dwell = np.array(batchinfo.dwell[np.isfinite(batchinfo.dwell)])
            self.dt = np.array(batchinfo.dt[np.isfinite(batchinfo.dt)])
            startpoints = np.array(batchinfo.startpoints[np.isfinite(batchinfo.startpoints)])
            endpoints = np.array(batchinfo.endpoints[np.isfinite(batchinfo.endpoints)])
            
            for i,dwell in enumerate(self.dwell):
                print(str(i) + '/' + str(len(self.dwell)))
                toffset = (eventtime[-1] + .75*eventbuffer)/self.outputsamplerate
                if i < len(self.dt)-1 and dwell > self.mindwell and self.frac[i] >self.minfrac:
                    if endpoints[i]+eventbuffer>startpoints[i+1]:
                        print('overlapping event')
                    else:
                        eventdata = self.data[startpoints[i]-eventbuffer:endpoints[i]+eventbuffer]
                        eventtime = np.arange(0,len(eventdata)) + .75*eventbuffer + eventtime[-1]
                        self.p1.plot(eventtime/self.outputsamplerate, eventdata,pen='b')
                        cusum = detect_cusum(self, eventdata, basesd = np.std(eventdata[0:eventbuffer])
                            , dt = 1/self.outputsamplerate, threshhold  = cusumthresh
                            , stepsize = cusumstep, minlength = self.minlevelt*self.outputsamplerate, maxstates = 10)
                        
                        while len(cusum['CurrentLevels']) < 3:
                            cusumthresh = cusumthresh *.9
                            cusumstep = cusumstep * .9
                            cusum = detect_cusum(self, eventdata, basesd = np.std(eventdata[0:eventbuffer])
                                , dt = 1/self.outputsamplerate, threshhold  = cusumthresh
                                , stepsize = cusumstep, minlength = self.minlevelt*self.outputsamplerate, maxstates = 10)
                      
#                        print len(cusum['CurrentLevels'])

                        
#                        if np.max(cusum['CurrentLevels'])-np.min(cusum['CurrentLevels']) == 0:
#                            cusum = detect_cusum(self, eventdata, basesd = np.std(eventdata)
#                                , dt = 1/self.outputsamplerate, threshhold  = cusumthresh/10
#                                , stepsize = cusumstep/10, minlength = self.minlevelt*self.outputsamplerate)
                            
                        ll = np.concatenate((ll,[(np.max(cusum['CurrentLevels'])-np.min(cusum['CurrentLevels']))/np.max(cusum['CurrentLevels'])]))
                        cusumthresh = cusum['Threshold']
                        cusumstep = cusum['stepsize']
                                                    
                        
                        for j,level in enumerate(cusum['CurrentLevels']):
                            self.p1.plot(y = 2*[level], x = toffset + cusum['EventDelay'][j:j+2], pen = pg.mkPen( 'r', width = 5))
                            try:
                                self.p1.plot(y = cusum['CurrentLevels'][j:j+2], x = toffset + 2*[cusum['EventDelay'][j+1]], pen = pg.mkPen( 'r', width = 5))
                            except Exception:
                                pass

        np.savetxt(self.matfilename+'llDB.txt',ll,delimiter='\t')
        self.p1.autoRange()
        
        print('\007')
        
    
        
    def sizethepore(self):
        self.ps = PoreSizer()
        self.ps.show()

    

    
    def customCond(self):
        if self.ui.groupBox_5.isChecked():
            self.useCustomConductance = 1
            self.ui.conductanceText.setEnabled(False)
            self.ui.resistanceText.setEnabled(False)
            self.UpdateIV()
        else:
            self.useCustomConductance = 0
            self.ui.conductanceText.setEnabled(True)
            self.ui.resistanceText.setEnabled(True)
            self.UpdateIV()

    def SaveAllFigures(self):
        self.pp.close()

    def DisplaySettings(self):
        if self.ui.actionUse_Clipping.isChecked():
            self.p1.setClipToView(True)
            self.Plot()
        else:
            self.p1.setClipToView(False)
            self.Plot()
        if self.ui.actionUse_Downsampling.isChecked():
            self.p1.setDownsampling(ds=True, auto=True, mode='subsample')
            self.Plot()
        else:
            self.p1.setDownsampling(ds=False, auto=True, mode='subsample')
            self.Plot()

    def dkeypresses(self):
        if self.ui.eventplot.underMouse():
            # PDF to save images:
            filename = os.path.splitext(os.path.basename(self.datafilename))[0]
            dirname = os.path.dirname(self.datafilename)
            self.count = 1
            while os.path.isfile(dirname + os.sep + filename + '_AllSavedImages_' + str(self.count) + '.pdf'):
                self.count += 1
            self.pp = PdfPages(dirname + os.sep + filename + '_AllSavedImages_' + str(self.count) + '.pdf')
            uf.SaveDerivatives(self)

    def skeypressed(self):
        if self.ui.ivplot.underMouse():
            uf.MatplotLibIV(self)
        if self.ui.signalplot.underMouse():
            if not self.count:
                # PDF to save images:
                filename = os.path.splitext(os.path.basename(self.datafilename))[0]
                dirname = os.path.dirname(self.datafilename)
                self.count = 1
                while os.path.isfile(dirname + os.sep + filename + '_AllSavedImages_' + str(self.count) + '.pdf'):
                    self.count += 1
                self.pp = PdfPages(dirname + os.sep + filename + '_AllSavedImages_' + str(self.count) + '.pdf')
            print('Added to PDF_' + str(self.count))
            uf.MatplotLibCurrentSignal(self)
            self.ui.signalplot.setBackground('g')
            time.sleep(1)
            self.ui.signalplot.setBackground('w')
        if self.ui.eventplot.underMouse():
            print('Event Plot Saved...')
            uf.SaveEventPlotMatplot(self)

    def EventFiltering(self, who):
        if self.ui.actionPlot_Common_Events.isChecked():
            self.NumberOfEvents = len(self.CommonIndexes['i1'])
        if self.ui.actionPlot_i1_detected_only.isChecked():
            self.NumberOfEvents = len(self.OnlyIndexes['i1'])
        if self.ui.actionPlot_i2_detected_only.isChecked():
            self.NumberOfEvents = len(self.OnlyIndexes['i2'])

    def Test(self):
        print('Yeeeeaahhh')

def QCloseEvent(w):
    print('Application closed, config saved...')

def start():
    app = QtGui.QApplication(sys.argv)
    myapp = GUIForm()
    myapp.show()
    sys.exit(app.exec_())

if __name__ == "__main__":
    global myapp
    app = QtGui.QApplication(sys.argv)
    myapp = GUIForm()
    myapp.show()
    sys.exit(app.exec_())

